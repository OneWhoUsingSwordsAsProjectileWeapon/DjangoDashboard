from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Avg, Count, Sum, F, Case, When, IntegerField, DecimalField
from django.db.models.functions import TruncMonth, TruncWeek, TruncDay
from django.db import models
from django.http import JsonResponse, Http404, HttpResponse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from datetime import datetime, timedelta, date
from django.utils import timezone
import json
import calendar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import Listing, Booking, Review, ListingImage
from .forms import ListingForm, BookingForm, ReviewForm, ListingSearchForm, ListingImageForm
from notifications.tasks import send_email_notification
from subscriptions.services import SubscriptionService

class ListingListView(ListView):
    """View for displaying list of listings with search and filtering"""
    model = Listing
    template_name = 'listings/listing_list.html'
    context_object_name = 'listings'
    paginate_by = 12

    def get_queryset(self):
        queryset = Listing.objects.filter(
            is_active=True, 
            is_approved=True
        ).filter(
            Q(approval_record__status='approved') | Q(approval_record__isnull=True)
        )

        # Apply search and filters if form is submitted
        form = ListingSearchForm(self.request.GET)
        if form.is_valid():
            # Location search
            location = form.cleaned_data.get('location')
            if location:
                queryset = queryset.filter(
                    Q(city__icontains=location) | 
                    Q(state__icontains=location) |
                    Q(country__icontains=location)
                )

            # Date availability
            check_in = form.cleaned_data.get('check_in')
            check_out = form.cleaned_data.get('check_out')
            if check_in and check_out:
                # Get all bookings that overlap with requested dates
                overlapping_bookings = Booking.objects.filter(
                    Q(start_date__lte=check_out, end_date__gte=check_in),
                    status__in=['confirmed', 'pending']
                ).values_list('listing_id', flat=True)

                # Exclude listings with overlapping bookings
                queryset = queryset.exclude(id__in=overlapping_bookings)

            # Guest capacity
            guests = form.cleaned_data.get('guests')
            if guests:
                queryset = queryset.filter(accommodates__gte=guests)

            # Price range
            min_price = form.cleaned_data.get('min_price')
            if min_price:
                queryset = queryset.filter(price_per_night__gte=min_price)

            max_price = form.cleaned_data.get('max_price')
            if max_price:
                queryset = queryset.filter(price_per_night__lte=max_price)

            # Property type
            property_type = form.cleaned_data.get('property_type')
            if property_type:
                queryset = queryset.filter(property_type=property_type)

            # Bedrooms
            bedrooms = form.cleaned_data.get('bedrooms')
            if bedrooms:
                queryset = queryset.filter(bedrooms__gte=bedrooms)

            # Bathrooms
            bathrooms = form.cleaned_data.get('bathrooms')
            if bathrooms:
                queryset = queryset.filter(bathrooms__gte=bathrooms)

        # Annotate with average rating
        queryset = queryset.annotate(
            avg_rating=Avg('reviews__rating'),
            review_count=Count('reviews')
        )

        # Handle sorting
        sort_by = self.request.GET.get('sort')
        if sort_by == 'price_asc':
            queryset = queryset.order_by('price_per_night')
        elif sort_by == 'price_desc':
            queryset = queryset.order_by('-price_per_night')
        elif sort_by == 'rating':
            queryset = queryset.order_by('-avg_rating', '-review_count')
        elif sort_by == 'newest':
            queryset = queryset.order_by('-created_at')
        else:
            # Default sorting by creation date
            queryset = queryset.order_by('-created_at')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add search form to context
        context['form'] = ListingSearchForm(self.request.GET)

        # Add popular destinations based on bookings in last month
        context['popular_destinations'] = self.get_popular_destinations()
        return context

    def get_popular_destinations(self):
        """Get popular destinations based on booking count in last month"""
        from datetime import timedelta
        from django.utils import timezone
        from django.db.models import Count, Q

        # Get bookings from last month
        last_month = timezone.now() - timedelta(days=30)

        # Get cities with most bookings in last month
        popular_cities = Booking.objects.filter(
            created_at__gte=last_month,
            status__in=['confirmed', 'completed']
        ).values(
            'listing__city', 'listing__state', 'listing__country'
        ).annotate(
            booking_count=Count('id'),
            city_name=F('listing__city'),
            state_name=F('listing__state'),
            country_name=F('listing__country')
        ).filter(
            booking_count__gt=0
        ).order_by('-booking_count')[:8]

        # Add sample image for each destination
        destinations_with_images = []
        default_images = [
            "https://images.unsplash.com/photo-1571896349842-33c89424de2d?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80",
            "https://images.unsplash.com/photo-1551698618-1dfe5d97d256?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80", 
            "https://images.unsplash.com/photo-1549144511-f099e773c147?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80",
            "https://images.unsplash.com/photo-1581833971358-2c8b550f87b3?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80",
            "https://images.unsplash.com/photo-1520637836862-4d197d17c90a?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80",
            "https://images.unsplash.com/photo-1506905925346-21bda4d32df4?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80",
            "https://images.unsplash.com/photo-1519302959554-a75be0afc82a?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80",
            "https://images.unsplash.com/photo-1542640244-b5d31b9c4d06?ixlib=rb-4.0.3&auto=format&fit=crop&w=800&q=80"
        ]

        for idx, city_data in enumerate(popular_cities):
            destinations_with_images.append({
                'city': city_data['city_name'],
                'state': city_data['state_name'], 
                'country': city_data['country_name'],
                'booking_count': city_data['booking_count'],
                'image': default_images[idx % len(default_images)]
            })

        # If no bookings, return default popular destinations
        if not destinations_with_images:
            destinations_with_images = [
                {
                    'city': 'Чита',
                    'state': 'Забайкальский край',
                    'country': 'Россия',
                    'booking_count': 0,
                    'image': default_images[0]
                },
                {
                    'city': 'Сочи', 
                    'state': 'Краснодарский край',
                    'country': 'Россия',
                    'booking_count': 0,
                    'image': default_images[1]
                },
                {
                    'city': 'Казань',
                    'state': 'Татарстан',
                    'country': 'Россия', 
                    'booking_count': 0,
                    'image': default_images[2]
                },
                {
                    'city': 'Екатеринбург',
                    'state': 'Свердловская область',
                    'country': 'Россия',
                    'booking_count': 0,
                    'image': default_images[3]
                }
            ]

        return destinations_with_images

class ListingDetailView(DetailView):
    """View for displaying listing details"""
    model = Listing
    template_name = 'listings/listing_detail.html'
    context_object_name = 'listing'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        listing = self.get_object()

        # Add booking form
        context['booking_form'] = BookingForm(listing=listing)

        # Get reviews
        reviews = listing.reviews.filter(is_approved=True).select_related('reviewer')
        context['reviews'] = reviews

        # Check if user can leave a review
        user = self.request.user
        user_has_reviewed = False
        can_review = False

        if user.is_authenticated and user != listing.host:
            # Check if user has already reviewed this listing
            existing_review = Review.objects.filter(
                reviewer=user, 
                listing=listing
            ).first()

            user_has_reviewed = existing_review is not None
            context['user_has_reviewed'] = user_has_reviewed

            if not existing_review:
                context['review_form'] = ReviewForm()
                context['can_review'] = True
                # Find any completed booking for this user and listing
                completed_booking = Booking.objects.filter(
                    guest=user,
                    listing=listing,
                    status='completed'
                ).first()
                context['booking_for_review'] = completed_booking

        # Unavailable dates for calendar
        unavailable_dates = listing.get_unavailable_dates()
        context['unavailable_dates_json'] = json.dumps(unavailable_dates)
        context['unavailable_dates'] = unavailable_dates

        # Get similar listings
        similar_listings = self.get_similar_listings(listing)
        context['similar_listings'] = similar_listings

        return context

    def get_similar_listings(self, listing):
        """Get similar listings based on location, accommodates, and price range"""
        from django.db.models import Q, Case, When, IntegerField
        from decimal import Decimal

        # Calculate price range (±40% from current listing price) 
        price_min = listing.price_per_night * Decimal('0.6')
        price_max = listing.price_per_night * Decimal('1.4')

        # Get similar listings with scoring system
        similar = Listing.objects.filter(
            is_active=True,
            is_approved=True,
            approval_record__status='approved'
        ).exclude(
            id=listing.id  # Exclude current listing
        ).annotate(
            # Score based on similarity criteria
            similarity_score=Case(
                # Same city gets highest priority (score 100)
                When(city__iexact=listing.city, then=100),
                # Same state gets medium priority (score 50)
                When(state__iexact=listing.state, then=50),
                # Same country gets low priority (score 25)
                When(country__iexact=listing.country, then=25),
                default=0,
                output_field=IntegerField()
            ) + Case(
                # Same accommodates capacity (score 30)
                When(accommodates=listing.accommodates, then=30),
                # Similar capacity ±2 guests (score 15)
                When(
                    accommodates__gte=listing.accommodates - 2,
                    accommodates__lte=listing.accommodates + 2,
                    then=15
                ),
                default=0,
                output_field=IntegerField()
            ) + Case(
                # Similar price range (score 20)
                When(
                    price_per_night__gte=price_min,
                    price_per_night__lte=price_max,
                    then=20
                ),
                default=0,
                output_field=IntegerField()
            ),
            avg_rating=Avg('reviews__rating'),
            review_count=Count('reviews')
        ).filter(
            # At least some similarity (same city, state, or country + similar accommodates/price)
            Q(city__iexact=listing.city) |
            Q(state__iexact=listing.state) |
            (Q(country__iexact=listing.country) & (
                Q(accommodates__gte=listing.accommodates - 2, accommodates__lte=listing.accommodates + 2) |
                Q(price_per_night__gte=price_min, price_per_night__lte=price_max)
            ))
        ).order_by(
            '-similarity_score',  # Best match first
            '-avg_rating',        # Then by rating
            'price_per_night'     # Then by price
        )[:8]  # Limit to 8 similar listings

        return similar

class HostDashboardView(LoginRequiredMixin, TemplateView):
    """Comprehensive host dashboard view"""
    template_name = 'listings/host_dashboard.html'

    def get(self, request, *args, **kwargs):
        # Check if user is a host
        if not request.user.is_host:
            messages.error(request, "You are not registered as a host.")
            return redirect('listings:listing_create')

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Get filter parameters
        time_filter = self.request.GET.get('time_filter', '30')  # days
        listing_filter = self.request.GET.get('listing_filter', 'all')
        status_filter = self.request.GET.get('status_filter', 'all')

        # Custom date range
        custom_start = self.request.GET.get('custom_start')
        custom_end = self.request.GET.get('custom_end')

        # Calculate date range
        end_date = timezone.now().date()

        if custom_start and custom_end:
            try:
                start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
                end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
                time_filter = 'custom'
            except ValueError:
                start_date = end_date - timedelta(days=30)
        elif time_filter == '7':
            start_date = end_date - timedelta(days=7)
        elif time_filter == '30':
            start_date = end_date - timedelta(days=30)
        elif time_filter == '90':
            start_date = end_date - timedelta(days=90)
        elif time_filter == '365':
            start_date = end_date - timedelta(days=365)
        else:
            start_date = end_date - timedelta(days=30)

        # Get host's listings
        host_listings = Listing.objects.filter(host=user)

        # Apply listing filter
        if listing_filter != 'all' and listing_filter.isdigit():
            filtered_listings = host_listings.filter(id=listing_filter)
        else:
            filtered_listings = host_listings

        # Get host's bookings with filters
        host_bookings = Booking.objects.filter(
            listing__host=user,
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )

        # Apply listing filter to bookings
        if listing_filter != 'all' and listing_filter.isdigit():
            host_bookings = host_bookings.filter(listing_id=listing_filter)

        # Apply status filter
        if status_filter != 'all':
            host_bookings = host_bookings.filter(status=status_filter)

        # Calculate comprehensive statistics with filtered data
        all_bookings = Booking.objects.filter(listing__host=user)
        
        # Use filtered bookings for period-specific metrics
        period_bookings = host_bookings
        
        stats = {
            'total_listings': host_listings.count(),
            'active_listings': host_listings.filter(is_active=True).count(),
            'pending_listings': host_listings.filter(is_approved=False).count(),
            'total_bookings': period_bookings.count(),
            'pending_bookings': period_bookings.filter(status='pending').count(),
            'confirmed_bookings': period_bookings.filter(status='confirmed').count(),
            'completed_bookings': period_bookings.filter(status='completed').count(),
            'canceled_bookings': period_bookings.filter(status='canceled').count(),
            'total_revenue': all_bookings.filter(status='completed').aggregate(
                total=Sum('total_price')
            )['total'] or 0,
            'filtered_revenue': period_bookings.filter(status='completed').aggregate(
                total=Sum('total_price')
            )['total'] or 0,
            'average_booking_value': period_bookings.filter(status='completed').aggregate(
                avg=Avg('total_price')
            )['avg'] or 0,
            'total_guests': period_bookings.filter(status__in=['completed', 'confirmed']).aggregate(
                total=Sum('guests')
            )['total'] or 0,
        }

        # Calculate occupancy rate properly for filtered period
        total_possible_nights = 0
        total_booked_nights = 0

        # Apply listing filter if specified
        if listing_filter != 'all' and listing_filter.isdigit():
            active_listings = host_listings.filter(id=listing_filter, is_active=True, is_approved=True)
        else:
            active_listings = host_listings.filter(is_active=True, is_approved=True)

        for listing in active_listings:
            # Calculate available days since listing creation or start_date, whichever is later
            listing_created = listing.created_at.date() if listing.created_at else start_date
            listing_start = max(listing_created, start_date)
            listing_end = min(timezone.now().date(), end_date)

            # Only calculate if we have a valid period
            if listing_end > listing_start:
                days_available = (listing_end - listing_start).days
                if days_available > 0:
                    total_possible_nights += days_available

                    # Get bookings for this listing in the time period with status filter
                    listing_bookings = listing.bookings.filter(
                        status__in=['completed', 'confirmed'],
                        start_date__lt=listing_end,
                        end_date__gt=listing_start
                    )
                    
                    # Apply status filter if specified
                    if status_filter != 'all':
                        listing_bookings = listing_bookings.filter(status=status_filter)

                    for booking in listing_bookings:
                        # Calculate overlap between booking and our period
                        booking_start = max(booking.start_date, listing_start)
                        booking_end = min(booking.end_date, listing_end)

                        # Only count if there's actual overlap
                        if booking_end > booking_start:
                            nights = (booking_end - booking_start).days
                            if nights > 0:
                                total_booked_nights += nights

        # Calculate occupancy rate as percentage
        if total_possible_nights > 0:
            occupancy_percentage = (total_booked_nights / total_possible_nights) * 100
            stats['occupancy_rate'] = round(occupancy_percentage, 1)
        else:
            stats['occupancy_rate'] = 0

        stats['total_booked_nights'] = total_booked_nights
        stats['total_possible_nights'] = total_possible_nights

        # Revenue by month (last 12 months)
        from dateutil.relativedelta import relativedelta
        import calendar

        # Generate last 12 months data
        monthly_revenue = []
        current_date = timezone.now().date()

        month_names_ru = {
            1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
            7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
        }

        for i in range(11, -1, -1):  # Go from 11 months ago to current month
            target_date = current_date.replace(day=1) - relativedelta(months=i)
            month_end = target_date + relativedelta(months=1)

            # Use end_date for completed bookings (when they actually finished)
            month_revenue = all_bookings.filter(
                status='completed',
                end_date__gte=target_date,
                end_date__lt=month_end
            ).aggregate(
                total_revenue=Sum('total_price'),
                total_bookings=Count('id')
            )

            month_name = f"{month_names_ru[target_date.month]} {target_date.year}"

            monthly_revenue.append({
                'month': target_date.strftime('%Y-%m-01'),
                'month_name': month_name,
                'revenue': float(month_revenue['total_revenue'] or 0),
                'bookings': month_revenue['total_bookings'] or 0
            })

        # Quarterly revenue (last 8 quarters)
        quarterly_revenue = []
        # Start from current quarter and go back 8 quarters
        today = timezone.now().date()
        current_quarter_start = date(today.year, ((today.month - 1) // 3) * 3 + 1, 1)
        
        for i in range(7, -1, -1):  # Last 8 quarters
            quarter_start = current_quarter_start - relativedelta(months=i*3)
            quarter_end = quarter_start + relativedelta(months=3)

            quarter_revenue = all_bookings.filter(
                status='completed',
                end_date__gte=quarter_start,
                end_date__lt=quarter_end
            ).aggregate(
                total_revenue=Sum('total_price'),
                total_bookings=Count('id')
            )

            quarter_num = ((quarter_start.month - 1) // 3) + 1
            quarterly_revenue.append({
                'quarter': f"{quarter_start.year}-Q{quarter_num}",
                'quarter_name': f"Q{quarter_num} {quarter_start.year}",
                'revenue': float(quarter_revenue['total_revenue'] or 0),
                'bookings': quarter_revenue['total_bookings'] or 0
            })

        # Seasonal revenue analysis (by month across all years)
        seasonal_data = {}
        month_names_ru = [
            '', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
        ]

        # Get data for last 2 years to have meaningful seasonal analysis
        two_years_ago = timezone.now().date() - timedelta(days=730)
        
        for month in range(1, 13):
            month_bookings = all_bookings.filter(
                status='completed',
                end_date__month=month,
                end_date__gte=two_years_ago  # Only last 2 years for better seasonal patterns
            ).aggregate(
                total_revenue=Sum('total_price'),
                total_bookings=Count('id'),
                avg_revenue=Avg('total_price')
            )

            seasonal_data[month] = {
                'month': month,
                'month_name': month_names_ru[month],
                'total_revenue': float(month_bookings['total_revenue'] or 0),
                'bookings': month_bookings['total_bookings'] or 0,
                'avg_revenue': float(month_bookings['avg_revenue'] or 0)
            }

        seasonal_revenue = list(seasonal_data.values())

        # Bookings by status - use filtered bookings
        status_stats_qs = period_bookings.values('status').annotate(
            count=Count('id'),
            revenue=Sum(Case(
                When(status='completed', then='total_price'),
                default=0,
                output_field=DecimalField()
            ))
        )

        # Convert to list with proper formatting
        status_stats = []
        for item in status_stats_qs:
            if item['count'] > 0:  # Only include statuses that have bookings
                status_stats.append({
                    'status': item['status'],
                    'count': item['count'],
                    'revenue': float(item['revenue'] or 0)
                })

        # Top performing listings with proper annotations
        top_listings = host_listings.annotate(
            avg_rating=Avg('reviews__rating'),
            review_count=Count('reviews', distinct=True),
            bookings_count=Count('bookings', distinct=True),
            total_revenue=Sum(
                Case(
                    When(bookings__status='completed', then='bookings__total_price'),
                    default=0,
                    output_field=DecimalField()
                )
            )
        ).prefetch_related('bookings').order_by('-total_revenue')

        # Calculate occupancy days separately to avoid aggregation issues
        for listing in top_listings:
            total_nights = 0
            for booking in listing.bookings.filter(status__in=['completed', 'confirmed']):
                nights = (booking.end_date - booking.start_date).days
                total_nights += nights
            listing.occupancy_days = total_nights

        # Recent activity (last 10)
        recent_bookings = all_bookings.select_related(
            'listing', 'guest'
        ).order_by('-created_at')[:10]

        # Guest demographics
        guest_stats = all_bookings.filter(
            status__in=['completed', 'confirmed']
        ).aggregate(
            avg_guests_per_booking=Avg('guests'),
            max_guests=models.Max('guests'),
            min_guests=models.Min('guests')
        )

        # Booking trends (last 30 days)
        daily_bookings = all_bookings.filter(
            created_at__gte=timezone.now() - timedelta(days=30)
        ).annotate(
            day=TruncDay('created_at')
        ).values('day').annotate(
            count=Count('id')
        ).order_by('day')

        # Average response time to bookings
        pending_time = all_bookings.filter(
            status__in=['confirmed', 'canceled']
        ).annotate(
            response_time=F('updated_at') - F('created_at')
        ).aggregate(
            avg_response=Avg('response_time')
        )

        # Analytics data for charts
        today = timezone.now().date()
        last_30_days = today - timedelta(days=30)
        last_12_months = today - timedelta(days=365)

        # Get user's listings for filtering
        user_listings = self.request.user.listings.all()

        # Detailed listing statistics
        selected_listing_id = self.request.GET.get('detail_listing_id')
        custom_detail_start = self.request.GET.get('custom_detail_start')
        custom_detail_end = self.request.GET.get('custom_detail_end')

        listing_detail_stats = None
        
        if selected_listing_id and selected_listing_id.isdigit():
            try:
                selected_listing = host_listings.get(id=selected_listing_id)
                
                # Calculate date range for detailed stats
                current_date = timezone.now().date()
                
                if custom_detail_start and custom_detail_end:
                    try:
                        detail_start_date = datetime.strptime(custom_detail_start, '%Y-%m-%d').date()
                        detail_end_date = datetime.strptime(custom_detail_end, '%Y-%m-%d').date()
                    except ValueError:
                        # Fallback to last year if dates are invalid
                        detail_start_date = current_date.replace(year=current_date.year - 1, month=1, day=1)
                        detail_end_date = current_date
                else:
                    # Default: last 12 months from current date
                    detail_start_date = current_date.replace(year=current_date.year - 1, month=current_date.month, day=1)
                    detail_end_date = current_date

                # Get listing creation date - use listing start if it's later
                listing_created = selected_listing.created_at.date() if selected_listing.created_at else detail_start_date
                actual_start_date = max(detail_start_date, listing_created)
                actual_end_date = min(detail_end_date, current_date)

                # Always calculate stats even if period seems short
                total_possible_days = max(1, (actual_end_date - actual_start_date).days)
                
                # Используем ТОЧНО такую же логику как в общей статистике заполняемости
                # Считаем заполняемость так же как в главной статистике
                occupied_days = 0
                
                # Получаем все подтвержденные и завершенные бронирования для этого объявления за весь период
                occupancy_bookings = selected_listing.bookings.filter(
                    status__in=['completed', 'confirmed'],
                    start_date__lt=actual_end_date,
                    end_date__gt=actual_start_date
                )

                # Calculate occupied days - точно как в общей статистике
                for booking in occupancy_bookings:
                    # Calculate overlap between booking and our period
                    booking_start = max(booking.start_date, actual_start_date)
                    booking_end = min(booking.end_date, actual_end_date)
                    
                    # Only count if there's actual overlap
                    if booking_end > booking_start:
                        nights = (booking_end - booking_start).days
                        if nights > 0:
                            occupied_days += nights

                # Calculate occupancy rate
                occupancy_rate = (occupied_days / total_possible_days * 100) if total_possible_days > 0 else 0

                # Get revenue for the period - используем end_date для завершенных бронирований
                period_revenue = selected_listing.bookings.filter(
                    status='completed',
                    end_date__gte=actual_start_date,
                    end_date__lt=actual_end_date + timedelta(days=1)
                ).aggregate(
                    total_revenue=Sum('total_price')
                )['total_revenue'] or 0

                # Get booking stats - считаем все бронирования в периоде по created_at
                all_bookings_in_period = selected_listing.bookings.filter(
                    created_at__date__gte=actual_start_date,
                    created_at__date__lte=actual_end_date
                )
                period_booking_count = all_bookings_in_period.count()
                completed_bookings = all_bookings_in_period.filter(status='completed').count()
                
                # Calculate average booking value
                avg_booking_value = (period_revenue / completed_bookings) if completed_bookings > 0 else 0

                # Get monthly breakdown
                monthly_breakdown = []
                current_month = actual_start_date.replace(day=1)
                
                month_names_ru = {
                    1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
                    7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
                }

                while current_month <= actual_end_date:
                    # Calculate month end
                    if current_month.month == 12:
                        month_end = current_month.replace(year=current_month.year + 1, month=1)
                    else:
                        month_end = current_month.replace(month=current_month.month + 1)
                    
                    month_end = min(month_end, actual_end_date + timedelta(days=1))
                    month_actual_end = min(month_end - timedelta(days=1), actual_end_date)

                    # Days in month within our period
                    month_days = max(0, (month_actual_end - max(current_month, actual_start_date)).days + 1) if month_actual_end >= max(current_month, actual_start_date) else 0

                    # Get bookings for this month - используем ту же логику что и выше
                    month_occupancy_bookings = selected_listing.bookings.filter(
                        status__in=['completed', 'confirmed'],
                        start_date__lt=month_end,
                        end_date__gt=current_month
                    )

                    # Calculate occupied days for this month - точно как в общей логике
                    month_occupied = 0
                    for booking in month_occupancy_bookings:
                        booking_start = max(booking.start_date, max(current_month, actual_start_date))
                        booking_end = min(booking.end_date, month_actual_end)
                        
                        # Only count if there's actual overlap
                        if booking_end > booking_start:
                            nights = (booking_end - booking_start).days
                            if nights > 0:
                                month_occupied += nights

                    month_occupancy = (month_occupied / month_days * 100) if month_days > 0 else 0

                    # Revenue for this month - используем end_date как в общей статистике
                    month_revenue = selected_listing.bookings.filter(
                        status='completed',
                        end_date__gte=current_month,
                        end_date__lt=month_end
                    ).aggregate(total=Sum('total_price'))['total'] or 0
                    
                    # Bookings count - используем created_at для подсчета бронирований
                    month_bookings_count = selected_listing.bookings.filter(
                        created_at__date__gte=current_month,
                        created_at__date__lt=month_end.date() if hasattr(month_end, 'date') else month_end
                    ).count()

                    monthly_breakdown.append({
                        'month': current_month.strftime('%Y-%m'),
                        'month_name': f"{month_names_ru[current_month.month]} {current_month.year}",
                        'days': month_days,
                        'occupied_days': month_occupied,
                        'occupancy_rate': round(month_occupancy, 1),
                        'revenue': float(month_revenue),
                        'bookings': month_bookings_count
                    })

                    # Move to next month
                    if current_month.month == 12:
                        current_month = current_month.replace(year=current_month.year + 1, month=1)
                    else:
                        current_month = current_month.replace(month=current_month.month + 1)

                listing_detail_stats = {
                    'listing': selected_listing,
                    'period_start': actual_start_date,
                    'period_end': actual_end_date,
                    'total_possible_days': total_possible_days,
                    'occupied_days': occupied_days,
                    'occupancy_rate': round(occupancy_rate, 1),
                    'total_revenue': float(period_revenue),
                    'total_bookings': period_booking_count,
                    'completed_bookings': completed_bookings,
                    'avg_booking_value': round(float(avg_booking_value), 2),
                    'monthly_breakdown': monthly_breakdown
                }

            except Listing.DoesNotExist:
                listing_detail_stats = {
                    'error': 'Выбранное объявление не найдено'
                }
        elif selected_listing_id:
            # Show error only if listing_id was provided but invalid
            listing_detail_stats = {
                'error': 'Неверный ID объявления'
            }

        context.update({
            'stats': stats,
            'monthly_revenue': monthly_revenue,
            'quarterly_revenue': quarterly_revenue,
            'seasonal_revenue': seasonal_revenue,
            'status_stats': status_stats,
            'top_listings': top_listings,
            'recent_bookings': recent_bookings,
            'guest_stats': guest_stats,
            'daily_bookings': list(daily_bookings),
            'pending_time': pending_time,
            'host_listings_for_filter': host_listings,
            'current_filters': {
                'time_filter': time_filter,
                'listing_filter': listing_filter,
                'status_filter': status_filter,
                'custom_start': custom_start,
                'custom_end': custom_end,
            },
            'time_filter_options': [
                ('7', 'Последние 7 дней'),
                ('30', 'Последние 30 дней'),
                ('90', 'Последние 3 месяца'),
                ('365', 'Последний год'),
                ('custom', 'Выбрать период'),
            ],
            'status_filter_options': [
                ('all', 'Все статусы'),
                ('pending', 'В ожидании'),
                ('confirmed', 'Подтверждено'),
                ('completed', 'Завершено'),
                ('canceled', 'Отменено'),
            ],
            'start_date': start_date,
            'end_date': end_date,
            'listing_detail_stats': listing_detail_stats,
            'selected_listing_id': selected_listing_id,
            'custom_detail_start': custom_detail_start,
            'custom_detail_end': custom_detail_end
        })

        return context

class HostListingListView(LoginRequiredMixin, ListView):
    """View for host to see their own listings"""
    model = Listing
    template_name = 'listings/host_listings.html'
    context_object_name = 'listings'

    def get_queryset(self):
        # Only show listings owned by current user
        queryset = Listing.objects.filter(host=self.request.user)
        
        # If this is for profile view, show only active and approved listings
        if self.request.GET.get('profile_view'):
            queryset = queryset.filter(
                is_active=True,
                is_approved=True
            ).filter(
                Q(approval_record__status='approved') | Q(approval_record__isnull=True)
            )
            
        return queryset

    def get_template_names(self):
        # Return different template for tab content
        if self.request.GET.get('tab_content'):
            return ['listings/partials/host_listings_content.html']
        return [self.template_name]

class ListingCreateView(LoginRequiredMixin, CreateView):
    """View for creating a new listing"""
    model = Listing
    form_class = ListingForm
    template_name = 'listings/listing_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add subscription limits info
        limits = SubscriptionService.get_ads_limits(self.request.user.id)
        context['subscription_limits'] = limits

        return context

    def form_valid(self, form):
        """Process the valid form."""
        # Set the host to the current user
        form.instance.host = self.request.user

        # Check if user can create a new listing
        can_create, error_message = SubscriptionService.can_create_ad(self.request.user.id)

        if not can_create:
            messages.warning(self.request, error_message + " Выберите подходящий план подписки для продолжения.")
            return redirect('subscriptions:plan_list')  # Redirect to subscription plans

        # Automatically make user a host when they create their first listing
        if not self.request.user.is_host:
            self.request.user.is_host = True
            self.request.user.save()

        # Save the listing
        response = super().form_valid(form)

        messages.success(self.request, 'Объявление успешно создано и отправлено на модерацию. Теперь вы являетесь хостом!')
        return response

class ListingUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """View for updating an existing listing"""
    model = Listing
    form_class = ListingForm
    template_name = 'listings/listing_form.html'

    def test_func(self):
        # Check if current user is the host
        listing = self.get_object()
        return self.request.user == listing.host

    def dispatch(self, request, *args, **kwargs):
        # Check if user is banned
        from moderation.models import BannedUser
        try:
            ban = BannedUser.objects.get(user=request.user)
            if ban.is_active:
                messages.error(request, 'Ваша учетная запись заблокирована. Вы не можете редактировать объявления.')
                return redirect('listings:host_listings')
        except BannedUser.DoesNotExist:
            pass
        
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        # Reset approval status if significant fields changed
        significant_fields = ['price_per_night', 'title', 'description', 'address']
        if any(form.cleaned_data.get(field) != getattr(self.get_object(), field) for field in significant_fields):
            form.instance.is_approved = False

        response = super().form_valid(form)
        messages.success(self.request, 'Your listing has been updated.')
        return response

class ListingDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """View for deleting a listing"""
    model = Listing
    template_name = 'listings/listing_confirm_delete.html'
    success_url = reverse_lazy('listings:host_listings')

    def test_func(self):
        # Check if current user is the host
        listing = self.get_object()
        return self.request.user == listing.host

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        listing = self.get_object()
        
        # Check for active bookings
        active_bookings = listing.bookings.filter(
            status__in=['pending', 'confirmed']
        ).select_related('guest')
        
        context['active_bookings'] = active_bookings
        context['can_delete'] = not active_bookings.exists()
        
        return context

    def delete(self, request, *args, **kwargs):
        listing = self.get_object()
        
        # Check if there are active bookings
        active_bookings = listing.bookings.filter(
            status__in=['pending', 'confirmed']
        )
        
        if active_bookings.exists():
            messages.error(
                request, 
                f'Нельзя удалить объявление с активными бронированиями. '
                f'У вас есть {active_bookings.count()} активных бронирований. '
                f'Сначала отмените все активные бронирования или дождитесь их завершения.'
            )
            return redirect('listings:listing_detail', pk=listing.pk)
        
        messages.success(request, 'Объявление успешно удалено.')
        return super().delete(request, *args, **kwargs)

@login_required
def add_listing_image(request, pk):
    """View for adding an image to a listing"""
    listing = get_object_or_404(Listing, pk=pk, host=request.user)

    if request.method == 'POST':
        form = ListingImageForm(request.POST, request.FILES)
        if form.is_valid():
            image = form.save(commit=False)
            image.listing = listing

            # If this is the first image or marked as main
            if form.cleaned_data.get('is_main') or not listing.images.exists():
                # Unset other main images
                listing.images.update(is_main=False)
                image.is_main = True

            image.save()
            messages.success(request, 'Изображение добавлено успешно.')
            return redirect('listings:listing_images', pk=listing.pk)
    else:
        form = ListingImageForm()

    return render(request, 'listings/add_image.html', {
        'form': form,
        'listing': listing
    })

@login_required
def manage_listing_images(request, pk):
    """View for managing listing images"""
    listing = get_object_or_404(Listing, pk=pk, host=request.user)

    return render(request, 'listings/manage_images.html', {
        'listing': listing
    })

@login_required
def remove_listing_image(request, pk, index):
    """View for removing an image from a listing"""
    listing = get_object_or_404(Listing, pk=pk, host=request.user)

    try:
        image_urls = listing.image_urls
        if 0 <= index < len(image_urls):
            # Remove the image at the specified index
            del image_urls[index]
            listing.image_urls = image_urls
            listing.save(update_fields=['image_urls'])
            messages.success(request, 'Image removed successfully.')
        else:
            messages.error(request, 'Invalid image index.')
    except (TypeError, IndexError):
        messages.error(request, 'Error removing image.')

    return redirect('listings:listing_images', pk=listing.pk)

def create_booking(request, pk):
    """View for creating a booking"""
    listing = get_object_or_404(
        Listing.objects.filter(
            pk=pk, 
            is_active=True, 
            is_approved=True,
            approval_record__status='approved'
        )
    )
    
    # If user is not authenticated, save the URL and redirect to login
    if not request.user.is_authenticated:
        # Save the current URL with all query parameters for after login
        request.session['redirect_after_login'] = request.get_full_path()
        
        # Save form data if provided in GET parameters
        booking_data = {}
        if 'check_in' in request.GET:
            booking_data['check_in'] = request.GET['check_in']
        if 'check_out' in request.GET:
            booking_data['check_out'] = request.GET['check_out']
        if 'guests' in request.GET:
            booking_data['guests'] = request.GET['guests']
        
        if booking_data:
            request.session['booking_form_data'] = booking_data
        
        return redirect('users:login')

    # Check if user is trying to book their own listing
    if request.user == listing.host:
        # Return to listing detail with modal trigger
        return redirect(f'{reverse("listings:listing_detail", kwargs={"pk": listing.pk})}?show_self_booking_modal=1')

    if request.method == 'POST':
        form = BookingForm(listing, request.POST)
        if form.is_valid():
            # Create booking but don't save yet
            booking = form.save(commit=False)
            booking.listing = listing
            booking.guest = request.user

            # Calculate the price
            start_date = form.cleaned_data['start_date']
            end_date = form.cleaned_data['end_date']
            price_data = listing.calculate_price(start_date, end_date)

            booking.base_price = price_data['base_price']
            booking.cleaning_fee = price_data['cleaning_fee']
            booking.service_fee = price_data['service_fee']
            booking.total_price = price_data['total_price']

            # Save the booking
            booking.save()

            # Send notifications
            guest_message = f"Your booking for {listing.title} has been created. " \
                           f"Check-in: {start_date}, Check-out: {end_date}. " \
                           f"Total price: ₽{booking.total_price}"

            host_message = f"New booking request for your listing '{listing.title}'. " \
                          f"Guest: {request.user.get_full_name() or request.user.username}, " \
                          f"Check-in: {start_date}, Check-out: {end_date}."

            # Send email notifications
            send_email_notification(
                request.user.email,
                "Booking Confirmation",
                guest_message
            )

            send_email_notification(
                listing.host.email,
                "New Booking Request",
                host_message
            )

            # Create in-app notifications
            from notifications.tasks import create_notification
            
            # Notification for guest
            create_notification(
                user=request.user,
                notification_type='booking_created',
                title=f"Запрос на бронирование создан",
                message=f"Ваш запрос на бронирование жилья '{listing.title}' создан и отправлен хозяину. Ожидайте подтверждения.",
                listing=listing,
                booking=booking
            )
            
            # Notification for host
            create_notification(
                user=listing.host,
                notification_type='booking_received',
                title=f"Новый запрос на бронирование",
                message=f"Пользователь {request.user.get_full_name() or request.user.username} хочет забронировать ваше жилье '{listing.title}' с {start_date} по {end_date}.",
                listing=listing,
                booking=booking
            )

            messages.success(request, "Booking created successfully. Awaiting host confirmation.")
            return redirect('listings:booking_detail', reference=booking.booking_reference)
    else:
        # Pre-fill form with query parameters or session data if provided
        initial = {}
        
        # First check for saved booking data in session
        booking_form_data = request.session.get('booking_form_data')
        if booking_form_data:
            if 'check_in' in booking_form_data:
                try:
                    initial['start_date'] = datetime.strptime(booking_form_data['check_in'], '%Y-%m-%d').date()
                except ValueError:
                    pass
            if 'check_out' in booking_form_data:
                try:
                    initial['end_date'] = datetime.strptime(booking_form_data['check_out'], '%Y-%m-%d').date()
                except ValueError:
                    pass
            if 'guests' in booking_form_data:
                try:
                    initial['guests'] = int(booking_form_data['guests'])
                except ValueError:
                    pass
            
            # Clear the session data after using it
            del request.session['booking_form_data']
        else:
            # Fall back to query parameters
            if 'check_in' in request.GET:
                try:
                    initial['start_date'] = datetime.strptime(request.GET['check_in'], '%Y-%m-%d').date()
                except ValueError:
                    pass

            if 'check_out' in request.GET:
                try:
                    initial['end_date'] = datetime.strptime(request.GET['check_out'], '%Y-%m-%d').date()
                except ValueError:
                    pass

            if 'guests' in request.GET:
                try:
                    initial['guests'] = int(request.GET['guests'])
                except ValueError:
                    pass

        form = BookingForm(listing, initial=initial)

    return render(request, 'listings/booking_form.html', {
        'form': form,
        'listing': listing
    })

@login_required
def booking_detail(request, reference):
    """View for displaying booking details"""
    try:
        # Get booking and check permissions
        booking = Booking.objects.select_related('listing', 'guest', 'listing__host').get(
            booking_reference=reference
        )

        # Only allow access to guest or host
        if request.user != booking.guest and request.user != booking.listing.host:
            raise Http404("Booking not found")

        return render(request, 'listings/booking_detail.html', {
            'booking': booking
        })
    except Booking.DoesNotExist:
        raise Http404("Booking not found")

@login_required
def user_bookings(request):
    """View for displaying user's bookings"""
    # Get all bookings for the user
    bookings = Booking.objects.filter(guest=request.user).select_related('listing')

    # Filter by status if provided
    status = request.GET.get('status')
    if status and status in [choice[0] for choice in Booking.STATUS_CHOICES]:
        bookings = bookings.filter(status=status)

    # Check if this is a request for tab content only
    if request.GET.get('tab_content'):
        return render(request, 'listings/partials/user_bookings_content.html', {
            'bookings': bookings,
            'status_choices': Booking.STATUS_CHOICES,
            'current_status': status
        })

    return render(request, 'listings/user_bookings.html', {
        'bookings': bookings,
        'status_choices': Booking.STATUS_CHOICES,
        'current_status': status
    })

@login_required
def host_bookings(request):
    """View for displaying host's bookings"""
    # Check if user is a host
    if not request.user.is_host:
        messages.error(request, "You are not registered as a host.")
        return redirect('listings:listing_list')

    # Get all bookings for the host's listings
    bookings = Booking.objects.filter(
        listing__host=request.user
    ).select_related('listing', 'guest')

    # Filter by status if provided
    status = request.GET.get('status')
    if status and status in [choice[0] for choice in Booking.STATUS_CHOICES]:
        bookings = bookings.filter(status=status)

    # Filter by specific listing if provided
    listing_id = request.GET.get('listing')
    if listing_id:
        try:
            bookings = bookings.filter(listing_id=int(listing_id))
        except ValueError:
            pass

    # Apply sorting
    sort_by = request.GET.get('sort', '-created_at')
    valid_sort_fields = [
        'created_at', '-created_at',
        'start_date', '-start_date', 
        'end_date', '-end_date',
        'total_price', '-total_price',
        'status', '-status',
        'guests', '-guests',
        'guest__username', '-guest__username',
        'listing__title', '-listing__title'
    ]
    
    if sort_by in valid_sort_fields:
        bookings = bookings.order_by(sort_by)
    else:
        bookings = bookings.order_by('-created_at')

    # Get host's listings for filter dropdown
    host_listings = Listing.objects.filter(host=request.user).order_by('title')

    return render(request, 'listings/host_bookings.html', {
        'bookings': bookings,
        'status_choices': Booking.STATUS_CHOICES,
        'current_status': status,
        'host_listings': host_listings,
        'current_listing': listing_id,
        'current_sort': sort_by,
        'sort_options': [
            ('-created_at', 'Дата создания (новые)'),
            ('created_at', 'Дата создания (старые)'),
            ('-start_date', 'Дата заезда (поздние)'),
            ('start_date', 'Дата заезда (ранние)'),
            ('-total_price', 'Цена (по убыванию)'),
            ('total_price', 'Цена (по возрастанию)'),
            ('status', 'Статус (А-Я)'),
            ('-status', 'Статус (Я-А)'),
            ('guest__username', 'Гость (А-Я)'),
            ('-guest__username', 'Гость (Я-А)'),
            ('listing__title', 'Объявление (А-Я)'),
            ('-listing__title', 'Объявление (Я-А)'),
        ]
    })

@login_required
def update_booking_status(request, reference, status):
    """View for updating booking status"""
    if status not in [choice[0] for choice in Booking.STATUS_CHOICES]:
        messages.error(request, "Invalid booking status.")
        return redirect('listings:host_bookings')

    try:
        booking = Booking.objects.select_related('listing').get(booking_reference=reference)

        # Verify permissions
        if request.user != booking.listing.host and not (request.user == booking.guest and status == 'canceled'):
            messages.error(request, "You don't have permission to perform this action.")
            return redirect('listings:listing_list')

        # Prevent certain status changes
        if booking.status == 'completed' and status != 'completed':
            messages.error(request, "Cannot change status of a completed booking.")
            return redirect('listings:booking_detail', reference=reference)

        # Check cancellation policy for confirmed bookings
        if status == 'canceled' and booking.status == 'confirmed':
            if not booking.can_be_canceled():
                from django.utils import timezone
                days_until_checkin = (booking.start_date - timezone.now().date()).days
                if days_until_checkin < 2:
                    messages.error(request, f"Нельзя отменить подтвержденное бронирование менее чем за 2 дня до заезда. До заезда осталось: {days_until_checkin} дн.")
                else:
                    messages.error(request, "Это бронирование нельзя отменить.")
                return redirect('listings:booking_detail', reference=reference)

        # Update booking status
        booking.status = status
        booking.save(update_fields=['status'])

        # Send notification
        if status == 'confirmed':
            message = f"Your booking for {booking.listing.title} has been confirmed by the host. " \
                     f"Check-in: {booking.start_date}, Check-out: {booking.end_date}."

            send_email_notification(
                booking.guest.email,
                "Booking Confirmed",
                message
            )

            # Create in-app notification for guest
            from notifications.tasks import create_notification
            create_notification(
                user=booking.guest,
                notification_type='booking_confirmed',
                title=f"Бронирование подтверждено!",
                message=f"Ваше бронирование жилья '{booking.listing.title}' подтверждено хозяином. Заезд: {booking.start_date}, Выезд: {booking.end_date}.",
                listing=booking.listing,
                booking=booking
            )
        elif status == 'canceled':
            # Determine who canceled
            canceler = "host" if request.user == booking.listing.host else "you"

            # Send to guest
            if request.user == booking.listing.host:
                message = f"Your booking for {booking.listing.title} has been canceled by the host. " \
                         f"Check-in: {booking.start_date}, Check-out: {booking.end_date}."

                send_email_notification(
                    booking.guest.email,
                    "Booking Canceled",
                    message
                )

                # Create in-app notification for guest
                from notifications.tasks import create_notification
                create_notification(
                    user=booking.guest,
                    notification_type='booking_canceled',
                    title=f"Бронирование отклонено",
                    message=f"Ваш запрос на бронирование жилья '{booking.listing.title}' был отклонен хозяином.",
                    listing=booking.listing,
                    booking=booking
                )

            # Send to host
            if request.user == booking.guest:
                message = f"Booking for {booking.listing.title} has been canceled by the guest. " \
                         f"Guest: {booking.guest.get_full_name() or booking.guest.username}, " \
                         f"Check-in: {booking.start_date}, Check-out: {booking.end_date}."

                send_email_notification(
                    booking.listing.host.email,
                    "Booking Canceled",
                    message
                )

                # Create in-app notification for host
                from notifications.tasks import create_notification
                create_notification(
                    user=booking.listing.host,
                    notification_type='booking_canceled_by_guest',
                    title=f"Бронирование отменено гостем",
                    message=f"Пользователь {booking.guest.get_full_name() or booking.guest.username} отменил бронирование жилья '{booking.listing.title}'.",
                    listing=booking.listing,
                    booking=booking
                )

        # Send automatic message to chat if conversation exists
        from chat.models import Conversation, Message
        try:
            conversation = Conversation.objects.get(booking=booking)
            if status == 'confirmed':
                auto_message = f"✅ Booking has been approved! Your stay from {booking.start_date} to {booking.end_date} is confirmed."
            elif status == 'canceled':
                auto_message = f"❌ Booking request has been declined."
            else:
                auto_message = f"ℹ️ Booking status updated to {booking.get_status_display()}."

            # Create system message
            Message.objects.create(
                conversation=conversation,
                sender=request.user,
                content=auto_message
            )
        except Conversation.DoesNotExist:
            pass

        messages.success(request, f"Booking status updated to {status}.")

        # Check for redirect parameter
        redirect_to = request.POST.get('redirect_to')
        if redirect_to:
            return redirect(redirect_to)

        return redirect('listings:booking_detail', reference=reference)

    except Booking.DoesNotExist:
        raise Http404("Booking not found")

@login_required
def create_review(request, listing_id):
    """View for creating a review for a listing"""
    listing = get_object_or_404(Listing, id=listing_id)

    # Check if user is the host
    if request.user == listing.host:
        messages.error(request, "Вы не можете оставить отзыв на свое собственное объявление.")
        return redirect('listings:listing_detail', pk=listing.pk)

    # Check if review already exists
    if Review.objects.filter(reviewer=request.user, listing=listing).exists():
        messages.error(request, "Вы уже оставили отзыв на это объявление.")
        return redirect('listings:listing_detail', pk=listing.pk)

    # Check if user has completed booking for this listing
    completed_booking = Booking.objects.filter(
        guest=request.user,
        listing=listing,
        status='completed'
    ).first()

    if not completed_booking:
        # Проверяем, есть ли вообще какие-то бронирования у пользователя для этого объявления
        any_booking = Booking.objects.filter(
            guest=request.user,
            listing=listing
        ).exists()

        if any_booking:
            messages.warning(request, "Вы можете оставить отзыв только после завершенного пребывания в этом месте. Ваше бронирование еще не завершено.")
        else:
            messages.warning(request, "Вы можете оставить отзыв только после того, как останетесь в этом месте. У вас нет завершенных бронирований для данного объявления.")

        return redirect('listings:listing_detail', pk=listing.pk)

    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.listing = listing
            review.reviewer = request.user
            review.booking = completed_booking
            review.save()

            # Send notification to host
            from notifications.tasks import create_notification
            create_notification(
                user=listing.host,
                notification_type='review_received',
                title=f"Новый отзыв для {listing.title}",
                message=f"{request.user.get_full_name() or request.user.username} оставил {review.rating}-звездочный отзыв для вашего объявления.",
                listing=listing,
                review=review
            )

            messages.success(request, "Ваш отзыв был успешно отправлен!")
            return redirect('listings:listing_detail', pk=listing.pk)
    else:
        form = ReviewForm()

    return render(request, 'listings/review_form.html', {
        'form': form,
        'listing': listing
    })

@login_required
def create_host_review(request, host_id):
    """View for creating a review for a host based on completed bookings"""
    from django.contrib.auth import get_user_model
    from .models import HostReview
    User = get_user_model()
    
    host = get_object_or_404(User, id=host_id)

    # Check if user is trying to review themselves
    if request.user == host:
        messages.error(request, "Вы не можете оставить отзыв на себя.")
        return redirect('users:public_profile', user_id=host.id)

    # Find a completed booking where the user was a guest and the host was the owner
    completed_booking = Booking.objects.filter(
        guest=request.user,
        listing__host=host,
        status='completed'
    ).first()

    if not completed_booking:
        messages.warning(request, "Вы можете оставить отзыв о хозяине только после завершенного пребывания в его объявлении.")
        return redirect('users:public_profile', user_id=host.id)

    # Check if review already exists for this booking
    if HostReview.objects.filter(reviewer=request.user, host=host, booking=completed_booking).exists():
        messages.error(request, "Вы уже оставили отзыв о хозяине для этого бронирования.")
        return redirect('users:public_profile', user_id=host.id)

    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            host_review = HostReview.objects.create(
                host=host,
                reviewer=request.user,
                booking=completed_booking,
                rating=form.cleaned_data['rating'],
                comment=form.cleaned_data['comment']
            )

            # Send notification to host
            from notifications.tasks import create_notification
            create_notification(
                user=host,
                notification_type='review_received',
                title=f"Новый отзыв от {request.user.get_full_name() or request.user.username}",
                message=f"Вы получили {host_review.rating}-звездочный отзыв за ваше гостеприимство.",
                booking=completed_booking
            )

            messages.success(request, "Ваш отзыв о хозяине был успешно отправлен!")
            return redirect('users:public_profile', user_id=host.id)
    else:
        form = ReviewForm()

    return render(request, 'listings/review_form.html', {
        'form': form,
        'listing': completed_booking.listing,
        'host_review': True,
        'host': host
    })

@login_required
def edit_review(request, review_id):
    """View for editing a review"""
    review = get_object_or_404(Review, id=review_id)

    # Check if user can edit this review
    if request.user != review.reviewer:
        messages.error(request, "Вы можете редактировать только свои собственные отзывы.")
        return redirect('listings:listing_detail', pk=review.listing.pk)

    if request.method == 'POST':
        form = ReviewForm(request.POST, instance=review)
        if form.is_valid():
            review = form.save(commit=False)
            review.is_edited = True
            review.save()
            messages.success(request, "Ваш отзыв был успешно обновлен!")
            return redirect('listings:listing_detail', pk=review.listing.pk)
    else:
        form = ReviewForm(instance=review)

    return render(request, 'listings/review_form.html', {
        'form': form,
        'listing': review.listing,
        'review': review,
        'is_edit': True
    })

@login_required
def delete_review(request, review_id):
    """View for deleting a review"""
    review = get_object_or_404(Review, id=review_id)

    # Check if user can delete this review
    if request.user != review.reviewer:
        messages.error(request, "Вы можете удалять только свои собственные отзывы.")
        return redirect('listings:listing_detail', pk=review.listing.pk)

    if request.method == 'POST':
        listing_id = review.listing.pk
        review.delete()
        messages.success(request, "Ваш отзыв был успешно удален.")
        return redirect('listings:listing_detail', pk=listing_id)

    return render(request, 'listings/review_confirm_delete.html', {
        'review': review
    })

@login_required
def admin_delete_review(request, review_id):
    """Admin view for deleting any review"""
    # Check if user is admin or moderator
    if not (request.user.is_staff or request.user.is_superuser or 
            request.user.role in ['admin', 'moderator']):
        messages.error(request, "У вас нет прав для выполнения этого действия.")
        return redirect('listings:listing_list')

    review = get_object_or_404(Review, id=review_id)

    if request.method == 'POST':
        listing_id = review.listing.pk
        reviewer_name = review.reviewer.get_full_name() or review.reviewer.username

        # Log the admin action - использовать правильные поля модели
        from moderation.models import ModerationLog
        try:
            ModerationLog.objects.create(
                moderator=request.user,
                description=f"Удален отзыв пользователя {reviewer_name} для объявления {review.listing.title}"
            )
        except Exception:
            # Если модель ModerationLog имеет другие поля, просто пропускаем логирование
            pass

        review.delete()
        messages.success(request, f"Отзыв пользователя {reviewer_name} был удален администратором.")
        return redirect('listings:listing_detail', pk=listing_id)

    return render(request, 'listings/admin_review_confirm_delete.html', {
        'review': review
    })

def get_listing_calendar_data(request, pk):
    """API view for getting listing calendar data"""
    try:
        listing = Listing.objects.get(pk=pk)

        # Get unavailable dates
        unavailable_dates = listing.get_unavailable_dates()

        # Return as JSON
        return JsonResponse({
            'unavailable_dates': unavailable_dates
        })
    except Listing.DoesNotExist:
        return JsonResponse({'error': 'Listing not found'}, status=404)

def calculate_booking_price(request, pk):
    """API view for calculating booking price"""
    try:
        listing = Listing.objects.get(pk=pk)

        # Get dates from request
        start_date = request.GET.get('start_date') or request.GET.get('check_in')
        end_date = request.GET.get('end_date') or request.GET.get('check_out')

        if not start_date or not end_date:
            # Return empty pricing template if no dates provided
            return render(request, 'listings/partials/_pricing.html', {
                'listing': listing,
                'check_in': None,
                'check_out': None
            })

        # Calculate price
        try:
            price_data = listing.calculate_price(start_date, end_date)

            # Check availability
            is_available = listing.is_available(start_date, end_date)

            # Prepare context for template
            context = {
                'listing': listing,
                'check_in': start_date,
                'check_out': end_date,
                'nights': price_data['nights'],
                'base_price': price_data['base_price'],
                'base_price_per_night': listing.price_per_night,
                'cleaning_fee': price_data['cleaning_fee'],
                'service_fee': price_data['service_fee'],
                'total_price': price_data['total_price'],
                'is_available': is_available
            }

            return render(request, 'listings/partials/_pricing.html', context)

        except (ValueError, TypeError):
            # Return error in pricing template
            return render(request, 'listings/partials/_pricing.html', {
                'listing': listing,
                'check_in': start_date,
                'check_out': end_date,
                'error': 'Неверный формат даты. Используйте ГГГГ-ММ-ДД.'
            })

    except Listing.DoesNotExist:
        return render(request, 'listings/partials/_pricing.html', {
            'error': 'Объявление не найдено'
        })

@login_required
def toggle_listing_status(request, pk):
    """API view for toggling listing active status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        listing = Listing.objects.get(pk=pk, host=request.user)

        # Check if user is banned
        from moderation.models import BannedUser
        try:
            ban = BannedUser.objects.get(user=request.user)
            if ban.is_active:
                return JsonResponse({
                    'error': 'Ваша учетная запись заблокирована. Вы не можете активировать объявления.',
                    'success': False
                }, status=403)
        except BannedUser.DoesNotExist:
            pass

        # Prevent activation if trying to activate
        if not listing.is_active:  # User wants to activate
            try:
                ban = BannedUser.objects.get(user=request.user)
                if ban.is_active:
                    return JsonResponse({
                        'error': 'Ваша учетная запись заблокирована. Вы не можете активировать объявления.',
                        'success': False
                    }, status=403)
            except BannedUser.DoesNotExist:
                pass

        # Toggle the status
        listing.is_active = not listing.is_active
        listing.save(update_fields=['is_active'])

        return JsonResponse({
            'success': True,
            'is_active': listing.is_active,
            'message': f'Listing {"activated" if listing.is_active else "deactivated"} successfully.'
        })

    except Listing.DoesNotExist:
        return JsonResponse({'error': 'Listing not found'}, status=404)

@login_required
def remove_image_file(request, pk, image_id):
    """View for removing a file image from a listing"""
    listing = get_object_or_404(Listing, pk=pk, host=request.user)
    image = get_object_or_404(ListingImage, pk=image_id, listing=listing)

    # If removing main image, set another image as main
    if image.is_main:
        other_image = ListingImage.objects.filter(listing=listing).exclude(pk=image_id).first()
        if other_image:
            other_image.is_main = True
            other_image.save()

    image.delete()
    messages.success(request, 'Изображение удалено.')
    return redirect('listings:listing_images', pk=listing.pk)

@login_required  
def set_main_image(request, pk, image_id):
    """View for setting an image as main"""
    listing = get_object_or_404(Listing, pk=pk, host=request.user)
    image = get_object_or_404(ListingImage, pk=image_id, listing=listing)

    # Unset current main image
    ListingImage.objects.filter(listing=listing, is_main=True).update(is_main=False)

    # Set new main image
    image.is_main = True
    image.save()

    messages.success(request, 'Главное изображение обновлено.')
    return redirect('listings:listing_images', pk=listing.pk)

@login_required
def user_reviews(request):
    """View for displaying user's reviews (given and received)"""
    user = request.user
    
    # Get reviews written by user
    reviews_given = Review.objects.filter(
        reviewer=user
    ).select_related('listing').order_by('-created_at')
    
    # Get reviews received by user (for their listings)
    reviews_received = Review.objects.filter(
        listing__host=user
    ).select_related('reviewer', 'listing').order_by('-created_at')
    
    # Check if this is a request for tab content only
    if request.GET.get('tab_content'):
        return render(request, 'listings/partials/user_reviews_content.html', {
            'reviews_given': reviews_given,
            'reviews_received': reviews_received
        })
    
    return render(request, 'listings/user_reviews.html', {
        'reviews_given': reviews_given,
        'reviews_received': reviews_received
    })

@login_required
def host_dashboard_data(request):
    """AJAX endpoint for updating dashboard chart data"""
    if not request.user.is_host:
        return JsonResponse({'error': 'Not a host'}, status=403)
    
    chart_type = request.GET.get('chart_type')
    period = request.GET.get('period', '12')
    
    # Get current filter parameters from session or request
    time_filter = request.GET.get('time_filter', '30')
    listing_filter = request.GET.get('listing_filter', 'all')
    status_filter = request.GET.get('status_filter', 'all')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')
    
    user = request.user
    all_bookings = Booking.objects.filter(listing__host=user)
    
    # Calculate date range for filtering (same logic as main view)
    end_date = timezone.now().date()
    if custom_start and custom_end:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
        except ValueError:
            start_date = end_date - timedelta(days=30)
    elif time_filter == '7':
        start_date = end_date - timedelta(days=7)
    elif time_filter == '30':
        start_date = end_date - timedelta(days=30)
    elif time_filter == '90':
        start_date = end_date - timedelta(days=90)
    elif time_filter == '365':
        start_date = end_date - timedelta(days=365)
    else:
        start_date = end_date - timedelta(days=30)
    
    # Apply filters to bookings
    filtered_bookings = all_bookings.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    )
    
    if listing_filter != 'all' and listing_filter.isdigit():
        filtered_bookings = filtered_bookings.filter(listing_id=listing_filter)
    
    if status_filter != 'all':
        filtered_bookings = filtered_bookings.filter(status=status_filter)
    
    response_data = {}
    
    if chart_type == 'revenue':
        # Generate revenue data based on period
        from dateutil.relativedelta import relativedelta
        
        try:
            period_months = int(period) if period != 'all' else 12
        except ValueError:
            period_months = 12
            
        monthly_revenue = []
        current_date = timezone.now().date()
        
        month_names_ru = {
            1: 'Янв', 2: 'Фев', 3: 'Мар', 4: 'Апр', 5: 'Май', 6: 'Июн',
            7: 'Июл', 8: 'Авг', 9: 'Сен', 10: 'Окт', 11: 'Ноя', 12: 'Дек'
        }
        
        for i in range(period_months - 1, -1, -1):
            target_date = current_date.replace(day=1) - relativedelta(months=i)
            month_end = target_date + relativedelta(months=1)
            
            month_revenue = all_bookings.filter(
                status='completed',
                end_date__gte=target_date,
                end_date__lt=month_end
            ).aggregate(
                total_revenue=Sum('total_price'),
                total_bookings=Count('id')
            )
            
            month_name = f"{month_names_ru[target_date.month]} {target_date.year}"
            
            monthly_revenue.append({
                'month': target_date.strftime('%Y-%m-01'),
                'month_name': month_name,
                'revenue': float(month_revenue['total_revenue'] or 0),
                'bookings': month_revenue['total_bookings'] or 0
            })
        
        response_data['monthly_revenue'] = monthly_revenue
    
    elif chart_type == 'quarterly':
        # Generate quarterly data
        try:
            period_quarters = int(period) if period != 'all' else 8
        except ValueError:
            period_quarters = 8
            
        quarterly_revenue = []
        today = timezone.now().date()
        current_quarter_start = date(today.year, ((today.month - 1) // 3) * 3 + 1, 1)
        
        for i in range(period_quarters - 1, -1, -1):
            quarter_start = current_quarter_start - relativedelta(months=i*3)
            quarter_end = quarter_start + relativedelta(months=3)
            
            quarter_revenue = all_bookings.filter(
                status='completed',
                end_date__gte=quarter_start,
                end_date__lt=quarter_end
            ).aggregate(
                total_revenue=Sum('total_price'),
                total_bookings=Count('id')
            )
            
            quarter_num = ((quarter_start.month - 1) // 3) + 1
            quarterly_revenue.append({
                'quarter': f"{quarter_start.year}-Q{quarter_num}",
                'quarter_name': f"Q{quarter_num} {quarter_start.year}",
                'revenue': float(quarter_revenue['total_revenue'] or 0),
                'bookings': quarter_revenue['total_bookings'] or 0
            })
        
        response_data['quarterly_revenue'] = quarterly_revenue
    
    elif chart_type == 'seasonal':
        # Generate seasonal data based on period
        seasonal_data = {}
        month_names_ru = [
            '', 'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
            'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
        ]
        
        # Filter by period
        if period == 'current':
            year_filter = timezone.now().year
            bookings_filter = all_bookings.filter(end_date__year=year_filter)
        else:
            # All time or last 2 years
            two_years_ago = timezone.now().date() - timedelta(days=730)
            bookings_filter = all_bookings.filter(end_date__gte=two_years_ago)
        
        for month in range(1, 13):
            month_bookings = bookings_filter.filter(
                status='completed',
                end_date__month=month
            ).aggregate(
                total_revenue=Sum('total_price'),
                total_bookings=Count('id'),
                avg_revenue=Avg('total_price')
            )
            
            seasonal_data[month] = {
                'month': month,
                'month_name': month_names_ru[month],
                'total_revenue': float(month_bookings['total_revenue'] or 0),
                'bookings': month_bookings['total_bookings'] or 0,
                'avg_revenue': float(month_bookings['avg_revenue'] or 0)
            }
        
        response_data['seasonal_revenue'] = list(seasonal_data.values())
    
    elif chart_type == 'status':
        # Generate status data using filtered bookings
        status_stats_qs = filtered_bookings.values('status').annotate(
            count=Count('id'),
            revenue=Sum(Case(
                When(status='completed', then='total_price'),
                default=0,
                output_field=DecimalField()
            ))
        )
        
        status_stats = []
        for item in status_stats_qs:
            if item['count'] > 0:
                status_stats.append({
                    'status': item['status'],
                    'count': item['count'],
                    'revenue': float(item['revenue'] or 0)
                })
        
        response_data['status_stats'] = status_stats
    
    return JsonResponse(response_data)

@login_required
def export_dashboard_excel(request):
    """Export dashboard data to Excel"""
    if not request.user.is_host:
        messages.error(request, "Вы не зарегистрированы как хост.")
        return redirect('listings:listing_list')

    user = request.user

    # Get filter parameters (same as dashboard)
    time_filter = request.GET.get('time_filter', '30')
    listing_filter = request.GET.get('listing_filter', 'all')
    status_filter = request.GET.get('status_filter', 'all')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')

    # Calculate date range
    end_date = timezone.now().date()

    if custom_start and custom_end:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
        except ValueError:
            start_date = end_date - timedelta(days=30)
    elif time_filter == '7':
        start_date = end_date - timedelta(days=7)
    elif time_filter == '30':
        start_date = end_date - timedelta(days=30)
    elif time_filter == '90':
        start_date = end_date - timedelta(days=90)
    elif time_filter == '365':
        start_date = end_date - timedelta(days=365)
    else:
        start_date = end_date - timedelta(days=30)

    # Get host's listings
    host_listings = Listing.objects.filter(host=user)
    if listing_filter != 'all' and listing_filter.isdigit():
        host_listings = host_listings.filter(id=listing_filter)

    # Get bookings
    bookings = Booking.objects.filter(
        listing__host=user,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('listing', 'guest')

    if listing_filter != 'all' and listing_filter.isdigit():
        bookings = bookings.filter(listing_id=listing_filter)
    if status_filter != 'all':
        bookings = bookings.filter(status=status_filter)

    # Create workbook
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create summary sheet
    summary_ws = wb.create_sheet(title="Сводка")

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # Summary data
    summary_data = [
        ["Период отчета", f"{start_date} - {end_date}"],
        ["", ""],
        ["Общая статистика", ""],
        ["Всего объявлений", host_listings.count()],
        ["Активных объявлений", host_listings.filter(is_active=True).count()],
        ["Всего бронирований", bookings.count()],
        ["Подтвержденных бронирований", bookings.filter(status='confirmed').count()],
        ["Завершенных бронирований", bookings.filter(status='completed').count()],
        ["Отмененных бронирований", bookings.filter(status='canceled').count()],
        ["Общий доход", bookings.filter(status='completed').aggregate(total=Sum('total_price'))['total'] or 0],
        ["Средний чек", bookings.filter(status='completed').aggregate(avg=Avg('total_price'))['avg'] or 0],
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=label)
        summary_ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 3:  # Header row
            summary_ws.cell(row=row_idx, column=1).font = header_font
            summary_ws.cell(row=row_idx, column=1).fill = header_fill

    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 20

    # Create bookings sheet
    bookings_ws = wb.create_sheet(title="Бронирования")

    # Bookings headers
    bookings_headers = [
        "ID бронирования", "Ссылка на бронирование", "Объявление", "Гость", 
        "Email гостя", "Дата заезда", "Дата выезда", "Количество ночей", 
        "Количество гостей", "Статус", "Базовая стоимость", "Сбор за уборку", 
        "Сервисный сбор", "Общая стоимость", "Дата создания", "Особые пожелания"
    ]

    # Write headers
    for col_idx, header in enumerate(bookings_headers, 1):
        cell = bookings_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write booking data
    for row_idx, booking in enumerate(bookings.order_by('-created_at'), 2):
        data = [
            str(booking.booking_reference),
            str(booking.booking_reference),
            str(booking.listing.title),
            str(booking.guest.get_full_name() or booking.guest.username),
            str(booking.guest.email),
            booking.start_date.strftime('%Y-%m-%d') if booking.start_date else '',
            booking.end_date.strftime('%Y-%m-%d') if booking.end_date else '',
            int(booking.duration_nights),
            int(booking.guests),
            str(booking.get_status_display()),
            float(booking.base_price),
            float(booking.cleaning_fee),
            float(booking.service_fee),
            float(booking.total_price),
            booking.created_at.strftime('%Y-%m-%d %H:%M'),
            str(booking.special_requests or "")
        ]

        for col_idx, value in enumerate(data, 1):
            bookings_ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in bookings_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        bookings_ws.column_dimensions[column_letter].width = adjusted_width

    # Create listings performance sheet
    listings_ws = wb.create_sheet(title="Эффективность объявлений")

    listings_headers = [
        "Название объявления", "Город", "Статус", "Цена за ночь",
        "Количество бронирований", "Завершенных бронирований", 
        "Общий доход", "Средний рейтинг", "Количество отзывов"
    ]

    # Write headers
    for col_idx, header in enumerate(listings_headers, 1):
        cell = listings_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Get listings with stats
    listings_with_stats = host_listings.annotate(
        total_bookings=Count('bookings', filter=Q(bookings__created_at__date__gte=start_date, bookings__created_at__date__lte=end_date)),
        completed_bookings=Count('bookings', filter=Q(bookings__status='completed', bookings__created_at__date__gte=start_date, bookings__created_at__date__lte=end_date)),
        total_revenue=Sum('bookings__total_price', filter=Q(bookings__status='completed', bookings__created_at__date__gte=start_date, bookings__created_at__date__lte=end_date)),
        avg_rating=Avg('reviews__rating'),
        review_count=Count('reviews')
    )

    # Write listings data
    for row_idx, listing in enumerate(listings_with_stats, 2):
        data = [
            listing.title,
            listing.city,
            "Активно" if listing.is_active else "Неактивно",
            float(listing.price_per_night),
            listing.total_bookings or 0,
            listing.completed_bookings or 0,
            float(listing.total_revenue or 0),
            round(listing.avg_rating or 0, 2),
            listing.review_count or 0
        ]

        for col_idx, value in enumerate(data, 1):
            listings_ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in listings_ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        listings_ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    filename = f"dashboard_report_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response